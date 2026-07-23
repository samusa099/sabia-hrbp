"""Generate Excel workbooks and preview images from the repository CSV files."""
from pathlib import Path
import csv
import sqlite3

ROOT = Path(__file__).resolve().parents[1]
CLEAN = ROOT / "06_Clean_Data"
MASTER = ROOT / "00_Master"
MASTER.mkdir(exist_ok=True)

def read_csv(name):
    path = CLEAN / name
    with path.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))

def to_number(value):
    if value is None or value == "":
        return None
    try:
        if "." in value:
            return float(value)
        return int(value)
    except (ValueError, TypeError):
        return value

def make_workbook(path, sheets):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.chart import LineChart, BarChart, Reference

    navy = "16324F"
    teal = "0F766E"
    light = "E8F0F7"
    white = "FFFFFF"
    border = Side(style="thin", color="D7DEE5")

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, title, rows in sheets:
        ws = wb.create_sheet(sheet_name[:31])
        headers = list(rows[0].keys()) if rows else ["Information"]
        end_col = max(1, len(headers))
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=end_col)
        cell = ws.cell(1, 1, title)
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color=white, bold=True, size=18)
        cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[1].height = 30
        for col, header in enumerate(headers, 1):
            c = ws.cell(4, col, header)
            c.fill = PatternFill("solid", fgColor=teal)
            c.font = Font(color=white, bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for r, row in enumerate(rows, 5):
            for c, header in enumerate(headers, 1):
                cell = ws.cell(r, c, to_number(row.get(header, "")))
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(left=border, right=border, top=border, bottom=border)
        ws.freeze_panes = "A5"
        for col in range(1, end_col + 1):
            letter = ws.cell(4, col).column_letter
            max_len = max([len(str(ws.cell(r, col).value or "")) for r in range(4, min(ws.max_row, 80)+1)] + [10])
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 34)

    wb.save(path)

employees = read_csv("employee_master.csv")
scorecard = read_csv("quarterly_scorecard.csv")
roadmap = read_csv("quarterly_roadmap.csv")
production = read_csv("production_monthly.csv")
finance = read_csv("financial_impact_monthly.csv")
recruitment = read_csv("recruitment_funnel.csv")
training = read_csv("training_records.csv")
performance = read_csv("performance_quarterly.csv")
attendance = read_csv("attendance_monthly.csv")
technology = read_csv("technology_adoption.csv")
reset = read_csv("layoff_and_restructuring.csv")
dictionary = read_csv("data_dictionary.csv")
pillars = read_csv("hr_pillars.csv")
pilot = read_csv("pilot_results.csv")

# Executive dashboard rows
dashboard = [
    {"KPI":"Q4 Active Headcount","Value":"114","Business Meaning":"Scaled workforce after a 48-role Q1 reset"},
    {"KPI":"Q4 Operating Profit","Value":"64.02 BDT million","Business Meaning":"Recovery moved into positive operating performance"},
    {"KPI":"Q4 Endpoint First-Pass Yield","Value":"97.1%","Business Meaning":"Precision manufacturing stabilized"},
    {"KPI":"Q4 Endpoint Defect Rate","Value":"2.4%","Business Meaning":"Quality loss sharply reduced"},
    {"KPI":"Q4 HRIS Adoption","Value":"97%","Business Meaning":"HR technology became an enterprise enabler"},
]
story = [
    {"Stage":"Business problem","Narrative":"Sabia entered smartwatch production without validating an MVP or prototype, creating high defect, rework, skill mismatch and operating loss."},
    {"Stage":"HRBP mandate","Narrative":"The Lead HRBP, also acting as Head of HR / Acting CHRO, connected workforce strategy, organization design, capability, HR services and technology with business recovery."},
    {"Stage":"Decision logic","Narrative":"The model scaled only after a controlled Q2 pilot outperformed the control group."},
]

make_workbook(
    MASTER / "Sabia_Group_HRBP_Analytics_Master_2026.xlsx",
    [
        ("Executive Dashboard","Sabia Group HRBP-Led Smartwatch Recovery",dashboard),
        ("Project Story","Problem Declaration and Business Recovery Narrative",story),
        ("Quarterly Roadmap","Q1-Q4 Approval-Gated Roadmap",roadmap),
        ("KPI Scorecard","Executive Quarterly KPI Scorecard",scorecard),
        ("HR Framework","HR Pillars and Operating Model",pillars),
        ("Employee Master","Employee Master and Skills Inventory",employees),
        ("Workforce Reset","Q1 Workforce Reset and Ethical Review",reset),
        ("Recruitment","Critical-Skill Recruitment Funnel",recruitment),
        ("Training","Learning and Critical-Skill Certification",training),
        ("Performance","Quarterly Performance Management",performance),
        ("Attendance","Department Attendance and Workload",attendance),
        ("Production","Smartwatch Production and Quality",production),
        ("Financial Impact","Monthly Benefits Realization",finance),
        ("Pilot Test","Q2 Controlled Pilot Results",pilot),
        ("HR Technology","HRIS Adoption and Data Quality",technology),
        ("Data Dictionary","Dataset Catalogue and Model Roles",dictionary),
    ],
)

quarters = {
    "Q1": ROOT / "01_Q1_Feasibility_and_Workforce_Reset",
    "Q2": ROOT / "02_Q2_Controlled_Pilot",
    "Q3": ROOT / "03_Q3_Scale_and_Stabilize",
    "Q4": ROOT / "04_Q4_Enterprise_Rollout",
}
for q, folder in quarters.items():
    folder.mkdir(exist_ok=True)
    rows = [r for r in roadmap if r.get("Quarter") == q]
    q_score = [{"Metric":r["Metric"],"Unit":r["Unit"],q:r[q],"Annual_Target":r["Annual_Target"]} for r in scorecard]
    extra = []
    if q == "Q1":
        extra = [("Workforce Reset","Q1 Role-Based Workforce Reset",reset)]
    elif q == "Q2":
        extra = [("Pilot Results","Q2 Pilot versus Control",pilot),("Training","Q2 Learning Records",[r for r in training if r.get("Quarter")=="Q2"])]
    elif q == "Q3":
        extra = [("Recruitment","Q3 Critical-Skill Hiring",[r for r in recruitment if r.get("Quarter")=="Q3"]),("Production","Q3 Production",[r for r in production if r.get("Quarter")=="Q3"])]
    else:
        extra = [("Financial Impact","Q4 Financial Outcomes",[r for r in finance if r.get("Quarter")=="Q4"]),("HR Technology","Q4 Technology Adoption",[r for r in technology if r.get("Quarter")=="Q4"])]
    make_workbook(folder / f"{q}_{folder.name.split('_',1)[1]}.xlsx", [("Cover",f"{q} Transformation Summary",rows),("Scorecard",f"{q} KPI Scorecard",q_score)] + extra)

# Preview images
import matplotlib.pyplot as plt
months = [r["Month"] for r in finance]
profits = [float(r["Operating_Profit_BDT_Million"]) for r in finance]
fig, ax = plt.subplots(figsize=(12,6))
ax.plot(months, profits, marker="o")
ax.axhline(0, linewidth=1)
ax.set_title("Sabia HRBP Recovery — Monthly Operating Profit")
ax.set_ylabel("BDT million")
ax.tick_params(axis="x", rotation=45)
fig.tight_layout()
fig.savefig(MASTER / "Executive_Dashboard_Preview.png", dpi=160)
plt.close(fig)

fig, ax = plt.subplots(figsize=(12,6))
names = ["Clean CSV tables","Raw staging tables","SQL reporting views","Quarterly workbooks"]
values = [27,3,15,4]
ax.bar(names, values)
ax.set_title("Database and Analytics Project Architecture")
ax.set_ylabel("Object count")
fig.tight_layout()
fig.savefig(MASTER / "Database_SQL_Sheet_Preview.png", dpi=160)
plt.close(fig)

print("Binary artifacts generated.")
